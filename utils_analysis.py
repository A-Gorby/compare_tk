import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment
from openpyxl import drawing

import matplotlib.pyplot as plt
# import seaborn as sns
# %matplotlib inline
from matplotlib.colors import ListedColormap, BoundaryNorm

from utils_enrichment import preprocess_tkbd
from utils_io import save_df_lst_to_excel
from utils_io import logger

if len(logger.handlers) > 1:
    for handler in logger.handlers:
        logger.removeHandler(handler)
    from utils_io import logger

from  local_dictionaries import sevice_sections, service_chapters, service_types_A, service_types_B, service_classes_A, service_classes_B
from  local_dictionaries import dict_ath_anatomy, dict_ath_therapy, dict_ath_pharm, dict_ath_chemical

def read_enriched_tk_data(path_tkbd_processed, fn_tk_bd):
    df_services = pd.read_excel(os.path.join(path_tkbd_processed, fn_tk_bd), sheet_name = 'Услуги')
    print(df_services.shape)
    display(df_services.head(2))
    df_LP = pd.read_excel(os.path.join(path_tkbd_processed, fn_tk_bd), sheet_name = 'ЛП')
    print(df_desc.shape)
    display(df_LP.head(2))
    df_RM = pd.read_excel(os.path.join(path_tkbd_processed, fn_tk_bd), sheet_name = 'РМ')
    print(df_RM.shape)
    display(df_RM.head(2))
    return df_services, df_LP, df_RM

    
def read_description(
    path_tk_models_source, fn_tk_description,
):
    df_desc = pd.read_excel(os.path.join(path_tk_models_source, fn_tk_description))
    print(df_desc.shape)
    display(df_desc.head(2))
    req_cols = ['Наименование ТК', 'Код ТК', 'Профиль', 'Наименование ТК (короткое)', 'Модель пациента', 'Файл Excel']

    if not set(req_cols).issubset(list(df_desc.columns)):
        logger.error(f"Файл описания моделей содержит неправильные колонки")
        logger.error(f"Файл описания моделей должен содержать колонки: {str(rec_cols)}")
        sys.exit(2)
    
    df_desc.duplicated(subset=['Наименование ТК'])
    tk_any_models = df_desc[df_desc.duplicated(subset=['Наименование ТК'])]['Наименование ТК'].values
    # tk_any_models
    tk_models = {}
    # tk_models = defaultdict()
    for tk_name in tk_any_models:
        mask_tk = df_desc['Наименование ТК']==tk_name
        models_desc = df_desc[mask_tk].values
        models = models_desc[:,3]
        # print(models)
        for i_row, row in df_desc[mask_tk].iterrows():
            # 'Профиль', 'Код ТК', 'Наименование ТК',
            if tk_models.get(tk_name) is None:
                tk_models[tk_name] = {}
                tk_models[tk_name]['Модели'] = []
                # tk_models.setdefault('tk_name', []).append(row['Наименование ТК'])
            tk_models[tk_name]['Код ТК'] = row['Код ТК']
            tk_models[tk_name]['Профиль'] = row['Профиль']
            tk_models[tk_name]['Наименование ТК (короткое)'] = row['Наименование ТК (короткое)']
            #tk_models[tk_name]['Модели'].append (dict(zip(['Модель пациента', 'Файл Excel',
           '#Название листа в файле Excel', 'Услуги', 'ЛП', 'РМ'], row.values[4:])))
           tk_models[tk_name]['Модели'].append (dict(zip(['Модель пациента', 'Файл Excel',], row.values[4:6])))
        # for model in models:
        # print(models_desc)
    return tk_models

def format_excel_cols_short(ws, format_cols, auto_filter=False):
    l_alignment=Alignment(horizontal='left', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    r_alignment=Alignment(horizontal='right', vertical= 'top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
    last_cell = ws.cell(row=1, column=len(format_cols)) 
    full_range = "A1:" + last_cell.column_letter + str(ws.max_row)
    if auto_filter:
        ws.auto_filter.ref = full_range
    ws.freeze_panes = ws['B2']
    for ic, col_width in enumerate(format_cols):
        cell = ws.cell(row=1, column=ic+1)
        cell.alignment = l_alignment
        ws.column_dimensions[cell.column_letter].width = col_width
    return ws


def change_order_base_techno(new_columns_02):
    if 'База' in new_columns_02 and 'Техно' in new_columns_02:
        i_base = new_columns_02.index('База')
        i_techno = new_columns_02.index('Техно')
        # print(i_base, i_techno)
        if i_techno < i_base:
            new_columns_03 = [col for col in new_columns_02 if col not in ['Техно', 'База']]
            # print(new_columns_03)
            if i_base > 0: i_base -= 1
            new_columns_03.insert(i_base, 'Техно')
            new_columns_03.insert(i_base, 'База')
            
            
            return new_columns_03
            
        else: return new_columns_02
    else: return new_columns_02


def simplify_multi_index (df_p, tk_names, model_names):
    '''
    on enter pdDataFrame with columns 
    MultiIndex([('count',  'Техкарта БА КС база.xlsx'), ('count', 'Техкарта БА КС техно.xlsx')], names=[None, 'Файл Excel'])
    '''
    pp_lst = []
    df_pp = df_p.reset_index()
    for i_row, row in df_pp.iterrows():
        pp_lst.append(row.values)
    # print(pp_lst[:2])
    cur_columns = list(df_pp.columns)
    # cur_columns: [('Код раздела', ''), ('count', 'Техкарта БА КС база.xlsx'), ('count', 'Техкарта БА КС техно.xlsx')]
    # print("cur_columns:", cur_columns)
    new_columns = [v[0] if i_v in [0,3] else v[1] for i_v, v in enumerate(cur_columns)]
    # print("new_columns:", new_columns)
    cur_columns_02 = list(df_pp.columns[1:3])
    # print("cur_columns_02:", cur_columns_02)
    # new_columns_02 = ['База' if (str(col[1])==tk_names[0]) else 'Техно' for col in cur_columns_02]
    new_columns_02 = [model_names[0] if (str(col[1])==tk_names[0]) else model_names[1] for col in cur_columns_02]
    new_columns_02 = [new_columns[0]] + new_columns_02 # + [new_columns[-1]] #+ code_names_columns 
    # print("new_columns_02:", new_columns_02)
    
    df_pp = pd.DataFrame(pp_lst, columns = new_columns_02)
    new_columns_03 = change_order_base_techno(new_columns_02)
    df_pp = df_pp[new_columns_03]
            
    return df_pp



def simplify_multi_index_02 (df_p, tk_names, model_names):
# def simpl_multi_index_02 (df_p, tk_names, model_names):    
    '''
    on enter pdDataFrame with columns 
    MultiIndex([('count',  'Техкарта БА КС база.xlsx'), ('count', 'Техкарта БА КС техно.xlsx')], names=[None, 'Файл Excel'])
    '''
    pp_lst = []
    df_pp = df_p.reset_index()
    for i_row, row in df_pp.iterrows():
        pp_lst.append(row.values)
    # print(pp_lst[:2])
    cur_columns = list(df_pp.columns)
    # print("cur_columns:", cur_columns)
    new_columns = [v[0] if i_v in [0] else v[-1] for i_v, v in enumerate(cur_columns)]
    # new_columns = [v[0]  for i_v, v in enumerate(cur_columns[:3])]
    # print("new_columns:", new_columns)
    # cur_columns_02 = list(df_pp.columns[1:3])
    # print("cur_columns_02:", cur_columns_02)
    # new_columns_02 = ['База' if (str(col[1])==tk_names[0]) else 'Техно' for col in cur_columns_02]
    new_columns_02 = [model_names[0] if (str(col)==tk_names[0]) else model_names[1] for col in new_columns[1:]]
    new_columns_02 = [new_columns[0]] + new_columns_02 
    # print("new_columns_02:", new_columns_02)
    
    df_pp = pd.DataFrame(pp_lst, columns = new_columns_02)
    new_columns_03 = change_order_base_techno(new_columns_02)
    df_pp = df_pp[new_columns_03]
            
    return df_pp

def def_differencies(df_pp, tk_names, model_names, code_names_columns, function_extract_names=None):
    diff_col_name = 'Разница'
    diff_lst = []
    try:
        df_pp[diff_col_name] = df_pp[[model_names[0], model_names[1]]].progress_apply(lambda x: pd.Series( x[0]-x[1]), axis=1)
    except Exception as err:
        print(err)
        df_pp[diff_col_name] = df_pp[[model_names[0], model_names[1]]].progress_apply(lambda x: pd.Series( float(x[0])-float(x[1])), axis=1)
    
    for i_row, row in df_pp[df_pp[diff_col_name]!=0].iterrows():
        if function_extract_names is not None:
            code_names = function_extract_names(row.values[0])
            # diff_lst.append([row.values[0], *[v for v in row.values[1:]],*[n for n in code_names] ])
            diff_lst.append([*[v for v in row.values],*[n for n in code_names] ])
        else:
            diff_lst.append(list(row.values))
    
    if function_extract_names is not None:
        new_columns = list(df_pp.columns) + code_names_columns
    else: 
        new_columns =  list(df_pp.columns)
    diff_df = pd.DataFrame(diff_lst, columns = new_columns)
    
    return diff_df


def extract_names_from_code_service(code, debug=False):
    section_name, type_name, class_name = None, None, None
    if (type(code)!= str) or ((type(code)==str) and (len(code)==0)): return section_name, type_name, class_name
    section_name = sevice_sections.get(code[0])
    if len(code)>=3:
        if code[0] == 'A':
            if debug: print('A')
            type_name = service_types_A.get(code[1:3])
        else: type_name = service_types_B.get(code[1:3])
        if len(code)>=6:
            if code[0] == 'A':
                class_name = service_classes_A.get(code[4:6])
            else: class_name = service_classes_B.get(code[4:7])
        else: return section_name, type_name, class_name
    else: return section_name, type_name, class_name
    
    return section_name, type_name, class_name

def services_comparison(
    df_services, tk_names, model_names,
    col_to_compare = 'Наименование услуги по Номенклатуре медицинских услуг (Приказ МЗ №804н)'):
    
    df1 = df_services[df_services['Файл Excel']==tk_names[0]]
    df2 = df_services[df_services['Файл Excel']==tk_names[1]]
    # print(df1.shape, df2.shape)
    print(f"Количество услуг: {model_names[0]}: {df1.shape}, {model_names[1]}: {df2.shape}")
    
    services_df1 = df1[col_to_compare].unique()
    services_df2 = df2[col_to_compare].unique()
    print(f"Количество уникальных услуг: {model_names[0]}: {len(services_df1)}, {model_names[1]}: {len(services_df2)}")
    
    common_services = [s for s in services_df1 if s in services_df2] + [s for s in services_df2 if s in services_df1]
    common_services = list(set(common_services))
    print("Количество общих услуг", len(common_services))
    
    diff_services_1_f_2 = [s for s in services_df1 if s not in services_df2]
    print(f"Количество отличающихся услуг: {model_names[0]}: {len(diff_services_1_f_2)}")
    # print(f"{diff_services_1_f_2[:5]}")
    diff_services_2_f_1 = [s for s in services_df2 if s not in services_df1]
    print(f"Количество отличающихся услуг: {model_names[1]}: {len(diff_services_2_f_1)}")
    # print(f"{diff_services_2_f_1[:5]}")
    
    df_common_services = pd.DataFrame([['Общие', s] for s in common_services], columns=['Тип сравнения', 'Услуга'])
    # df_common_services.head(2)
    # print(df_common_services.shape)
    # pprint(diff_services_1_f_2)
    df_diff_services_1_f_2 = pd.DataFrame([[f"Есть в '{model_names[0]}', нет в '{model_names[1]}'", s] 
                                           for s in diff_services_1_f_2], columns=['Тип сравнения', 'Услуга'])
    # df_diff_services_1_f_2.head(2)
    df_diff_services_2_f_1 = pd.DataFrame([[f"Есть в '{model_names[1]}', нет в '{model_names[0]}'", s] 
                                           for s in diff_services_2_f_1], columns=['Тип сравнения', 'Услуга'])
    # df_diff_services_2_f_1.head(2)
    # print(df_diff_services_2_f_1.shape)
    df_services_compare = pd.concat([df_common_services, df_diff_services_1_f_2, df_diff_services_2_f_1])
    print(f"Услуги: Итого строк сравнения: {df_services_compare.shape[0]}")
    
    return df_services_compare

def LP_comparison(
    df_LP, tk_names, model_names,
    col_to_compare = 'Наименование лекарственного препарата (ЛП) (МНН)'):
    df1 = df_LP[df_LP['Файл Excel']==tk_names[0]]
    df2 = df_LP[df_LP['Файл Excel']==tk_names[1]]
    # print(df1.shape, df2.shape)
    print(f"Количество ЛП: {model_names[0]}: {df1.shape[0]}, {model_names[1]}: {df2.shape[0]}")
    LP_df1 = df1[col_to_compare].unique()
    LP_df2 = df2[col_to_compare].unique()
    print(f"Количество уникальных ЛП: {model_names[0]}: {len(LP_df1)}, {model_names[1]}: {len(LP_df2)}")
    
    common_LP = [s for s in LP_df1 if s in LP_df2] + [s for s in LP_df2 if s in LP_df1]
    common_LP = list(set(common_LP))
    print(f"Количество общих ЛП: {len(common_LP)}")

    diff_LP_1_f_2 = [s for s in LP_df1 if s not in LP_df2]
    print(f"Количество отличающихся ЛП: {model_names[0]}: {len(diff_LP_1_f_2)}")
    diff_LP_2_f_1 = [s for s in LP_df2 if s not in LP_df1]
    print(f"Количество отличающихся ЛП: {model_names[1]}: {len(diff_LP_2_f_1)}")
    
    df_common_LP = pd.DataFrame([['Общие ЛП(МНН)', s] for s in common_LP], columns=['Тип сравнения', 'ЛП(МНН)'])
    # df_common_LP.head(2)
    # print(df_common_LP.shape)
    # pprint(diff_services_1_f_2)
    df_diff_LP_1_f_2 = pd.DataFrame([[f"Есть в '{model_names[0]}', нет в '{model_names[1]}'", s] 
                                           for s in diff_LP_1_f_2], columns=['Тип сравнения', 'ЛП(МНН)'])
    # df_diff_services_1_f_2.head(2)
    df_diff_LP_2_f_1 = pd.DataFrame([[f"Есть в '{model_names[1]}', нет в '{model_names[0]}'", s] 
                                           for s in diff_LP_2_f_1], columns=['Тип сравнения', 'ЛП(МНН)'])
    
    df_LP_compare = pd.concat([df_common_LP, df_diff_LP_1_f_2, df_diff_LP_2_f_1])
    # print(df_LP_compare.shape)
    print(f"ЛП: Итого строк сравнения: {df_LP_compare.shape[0]}")
    
    return df_LP_compare

def RM_comparison(
    df_RM, tk_names, model_names,
    col_to_compare = 'Изделия медицинского назначения и расходные материалы, обязательно используемые при оказании медицинской услуги'):
    
    df1 = df_RM[df_RM['Файл Excel']==tk_names[0]]
    df2 = df_RM[df_RM['Файл Excel']==tk_names[1]]
    print(f"Количество МИ/РМ: {model_names[0]}: {df1.shape[0]}, {model_names[1]}: {df2.shape[0]}")
    RM_df1 = df1[col_to_compare].unique()
    RM_df2 = df2[col_to_compare].unique()
    print(f"Количество уникальных МИ/РМ: {model_names[0]}: {len(RM_df1)}, {model_names[1]}: {len(RM_df2)}")
    
    common_RM = [s for s in RM_df1 if s in RM_df2] + [s for s in RM_df2 if s in RM_df1]
    common_RM = list(set(common_RM))
    print(f"Количество общих МИ/РМ: {len(common_RM)}")
    
    diff_RM_1_f_2 = [s for s in RM_df1 if s not in RM_df2]
    print(f"Количество отличающихся МИ/РМ: {model_names[0]}: {len(diff_RM_1_f_2)}")
    diff_RM_2_f_1 = [s for s in RM_df2 if s not in RM_df1]
    print(f"Количество отличающихся МИ/РМ: {model_names[1]}: {len(diff_RM_2_f_1)}")
    
    df_common_RM = pd.DataFrame([['Общие МИ/РМ', s] for s in common_RM], columns=['Тип сравнения', 'МИ/РМ'])
    df_diff_RM_1_f_2 = pd.DataFrame([[f"Есть в '{model_names[0]}', нет в '{model_names[1]}'", s] 
                                           for s in diff_RM_1_f_2], columns=['Тип сравнения', 'МИ/РМ'])
    # df_diff_services_1_f_2.head(2)
    df_diff_RM_2_f_1 = pd.DataFrame([[f"Есть в '{model_names[1]}', нет в '{model_names[0]}'", s] 
                                           for s in diff_RM_2_f_1], columns=['Тип сравнения', 'МИ/РМ'])
    
    df_RM_compare = pd.concat([df_common_RM, df_diff_RM_1_f_2, df_diff_RM_2_f_1])
    # print(df_LP_compare.shape)
    print(f"ЛП: Итого строк сравнения: {df_RM_compare.shape[0]}")
    
    return df_RM_compare


def services_analysis(
    df_services, tk_names, model_names, tk_code_name,
    path_tk_models_processed
    ):
    
    codes_columns_services = ['Код раздела', 'Код типа', 'Код класса']
    code_names_columns_services = ['Раздел', 'Тип', 'Класс']
    services_mask_base = df_services['Файл Excel'] == tk_names[0]
    services_mask_techno = df_services['Файл Excel'] == tk_names[1]
    df_a = df_services[services_mask_base | services_mask_techno]
    # tk_name, model, analysis_part, analysis_part_code = 'Нейрохирургия',  'База', 'Услуги', 1
    tk_name, model, analysis_part, analysis_part_code = tk_code_name,  'База', 'Услуги', 1
    # dictionaries_lst = [sevice_sections, (service_types_A, service_types_B), (service_classes_A, service_classes_B) ]
    diff_lst = []
    diff_df_services = []
    # code_names_columns_services = ['Раздел', 'Тип', 'Класс']
    n_bars_max_on_picture = 20
    # from matplotlib.colors import ListedColormap, BoundaryNorm
    colors=["#9b59b6", "#3498db", "#95a5a6", "#e74c3c", "#34495e", "#2ecc71"]
    cmap = ListedColormap(["#95a5a6", "#2ecc71"])

    try:
        # os.mkdir(os.path.join(path_tk_models_processed, tk_code_name))
        os.mkdir(os.path.join(path_tk_models_processed, tk_code_name, 'img'))
    except:
        pass

    for i_col, col_name in enumerate(codes_columns_services):
        diff_lst.append([])
        df_p = pd.DataFrame({'count' : df_a.groupby( ['Файл Excel', col_name] ).size()}).reset_index().pivot([col_name], ['Файл Excel'] )\
        .fillna(0)
        # print(df_p.columns)
        # display(df_p.head(2))
        df_pp = simplify_multi_index (df_p, tk_names, model_names)
        # display(df_pp.head(2))
        kind = 'bar' #'kde' #'area' #'bar'
        title = '\n'.join([tk_code_name, analysis_part]) #, col_name])
        y_lim_min = 0

        # print("df_pp.shape[0]:", df_pp.shape[0])
        if df_pp.shape[0] <= n_bars_max_on_picture:
            plt.figure(figsize=(25, 6), tight_layout=True)
            ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)
        else:
            plt.figure(figsize=(25, 10), tight_layout=True)
            for i_max in range(10):
                # df_pp1 = df_pp[(df_pp['База']>=y_lim_min + i_max) | (df_pp['Техно']>=y_lim_min + i_max)]
                df_pp1 = df_pp[(df_pp[model_names[0]]>=y_lim_min + i_max) | (df_pp[model_names[1]]>=y_lim_min + i_max)]
                
                if df_pp1.shape[0] <= n_bars_max_on_picture:
                    ax1 = df_pp1.plot(kind= kind, x = col_name, rot=45, cmap = cmap) 
                    break

        legend_list = model_names
        ax1.legend(legend_list, loc='best',fontsize=8)
        plt.title(title, fontsize=8)
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)
        plt.xlabel(col_name, fontsize=8)
        plt.ylabel('Количество', fontsize=8)

        # fn_img = f"{analysis_part_code:02d}_{analysis_part}_{i_col:02d}.jpg"
        fn_img = f"{analysis_part_code:02d}_Услуги_01_Состав_{i_col:02d}.jpg" #.replace(' ','_')
        # plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, fn_img), bbox_inches='tight')
        plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, 'img', fn_img), bbox_inches='tight')
        plt.show()

        diff_df_services.append(def_differencies(
                             df_pp, tk_names, model_names,
                             code_names_columns = code_names_columns_services, 
                             function_extract_names = extract_names_from_code_service))
        display(diff_df_services[i_col])
    return diff_df_services

def services_analysis_02(
    df_services, tk_names, model_names, tk_code_name,
    path_tk_models_processed,
    analysis_subpart_code, analysis_subpart,
    indicator_col_name = 'Усредненная частота предоставления',
    agg_type = 'Среднее',
    
    ):
    
    codes_columns_services = ['Код раздела', 'Код типа', 'Код класса']
    code_names_columns_services = ['Раздел', 'Тип', 'Класс']
    services_mask_base = df_services['Файл Excel'] == tk_names[0]
    services_mask_techno = df_services['Файл Excel'] == tk_names[1]
    df_a = df_services[services_mask_base | services_mask_techno]
    # tk_name, model, analysis_part, analysis_part_code = 'Нейрохирургия',  'База', 'Услуги', 1
    # tk_name, model, analysis_part, analysis_part_code = tk_code_name,  'База', 'Услуги', 1
    # dictionaries_lst = [sevice_sections, (service_types_A, service_types_B), (service_classes_A, service_classes_B) ]
    diff_lst = []
    diff_df_services = []
    # code_names_columns_services = ['Раздел', 'Тип', 'Класс']
    n_bars_max_on_picture = 20
    # from matplotlib.colors import ListedColormap, BoundaryNorm
    colors=["#9b59b6", "#3498db", "#95a5a6", "#e74c3c", "#34495e", "#2ecc71"]
    cmap = ListedColormap(["#95a5a6", "#2ecc71"])

    for i_col, col_name in enumerate(codes_columns_services):
        diff_lst.append([])
        if agg_type == 'Среднее':
            df_p = df_a.groupby( ['Файл Excel', col_name, ] ).agg({indicator_col_name: ['mean']})\
                        .reset_index().pivot([col_name], ['Файл Excel'] ).fillna(0)
        elif agg_type == 'Сумма':
            df_p = df_a.groupby( ['Файл Excel', col_name, ] ).agg({indicator_col_name: ['sum']})\
                        .reset_index().pivot([col_name], ['Файл Excel'] ).fillna(0)
        # print(df_p.columns)
        # display(df_p.head(2))
        df_pp = simplify_multi_index_02 (df_p, tk_names, model_names)
        # df_pp = simpl_multi_index_02 (df_p, tk_names, model_names)
        # display(df_pp.head(2))
        kind = 'bar' #'kde' #'area' #'bar'
        title = '\n'.join([tk_code_name, 'Услуги', analysis_subpart]) #, indicator_col_name]) #, col_name])
        y_lim_min = 0

        # print("df_pp.shape[0]:", df_pp.shape[0])
        if df_pp.shape[0] <= n_bars_max_on_picture:
            plt.figure(figsize=(25, 6), tight_layout=True)
            ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)
        else:
            plt.figure(figsize=(25, 10), tight_layout=True)
            try:
                max_v = max(df_pp[model_names[0]].max(), df_pp[model_names[1]].max())
                min_v = min(df_pp[model_names[0]].min(), df_pp[model_names[1]].min())
                delta_v = (max_v - min_v)/10
                for i_max in range(10):
                    # df_pp1 = df_pp[(df_pp['База']>=y_lim_min + i_max) | (df_pp['Техно']>=y_lim_min + i_max)]

                    df_pp1 = df_pp[(df_pp[model_names[0]]>=y_lim_min + i_max*delta_v) | (df_pp[model_names[1]]>=y_lim_min + i_max*delta_v)]

                    if df_pp1.shape[0] <= n_bars_max_on_picture:
                        print(f"i_max: {i_max}, df_pp1.shape[0]: {df_pp1.shape[0]}")
                        ax1 = df_pp1.plot(kind= kind, x = col_name, rot=45, cmap = cmap) #, y_lim= (y_lim_min + i_max,100)) 
                        break
            except Exception as err:
                print(str(err))
                ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)

        legend_list = model_names
        ax1.legend(legend_list, loc='best',fontsize=8)
        plt.title(title, fontsize=8)
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)
        plt.xlabel(col_name, fontsize=8)
        # plt.ylabel('Количество', fontsize=8)
        plt.ylabel(agg_type, fontsize=8)

        # fn_img = f"{analysis_part_code:02d}_{analysis_part}_{i_col:02d}.jpg"
        fn_img = f"01_Услуги_{analysis_subpart_code:02d}_{analysis_subpart}_{i_col:02d}.jpg" #.replace(' ','_')
        
        # plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, fn_img), bbox_inches='tight')
        plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, 'img', fn_img), bbox_inches='tight')
        # plt.savefig(path_tk_models_processed + tk_code_name + '/' + fn_img, bbox_inches='tight')
        plt.show()
        try:
            diff_df_services.append(def_differencies(
                                 df_pp, tk_names, model_names,
                                 code_names_columns = code_names_columns_services, 
                                 function_extract_names = extract_names_from_code_service))
            display(diff_df_services[i_col])
        except Exception as err:
            diff_df_services.append(None)
            logger.error(str(err))
            logger.error(f"Данные анализа об отличиях не выводятся из-за некорректных входных данных")
    return diff_df_services

def extract_name_groups_ATH(s, debug = False):
    ath_anatomy_code, ath_anatomy, ath_therapy_code, ath_therapy, ath_pharm_code, ath_pharm, ath_chemical_code, ath_chemical = \
        None, None, None, None, None, None, None, None
    if type(s) is None or ((type(s)==float) and np.isnan(s)) or (type(s)!=str)  or ((type(s)==str) and (len(s)==0)):
        return None, None, None, None, None, None, None, None
    ath_anatomy_code = s[0]
    ath_anatomy = dict_ath_anatomy.get(ath_anatomy_code)
    if len(s)>=3:
        ath_therapy_code = s[0:3]
        ath_therapy = dict_ath_therapy.get(ath_therapy_code)
        if len(s)>=4:
            ath_pharm_code = s[0:4]
            ath_pharm = dict_ath_pharm.get(ath_pharm_code)
            if len(s)>=5:
                ath_chemical_code = s[0:5]
                ath_chemical = dict_ath_chemical.get(ath_chemical_code)
    return ath_anatomy, ath_therapy, ath_pharm, ath_chemical

def LP_analysis(
    df_LP, tk_names, model_names, tk_code_name,
    path_tk_models_processed
    ):
    
    lp_mask_base = df_LP['Файл Excel'] == tk_names[0]
    lp_mask_techno = df_LP['Файл Excel'] == tk_names[1]
    
    tk_name, analysis_part, analysis_part_code = tk_code_name, 'ЛП', 2
    columns_to_compare =['Код анатомического органа или системы',
       'Код терапевтической группы', 
       'Код фармакологической группы',
       'Код химической группы', 
       'Код группы ЛП (АТХ)', 
        'Форма выпуска лекарственного препарата (ЛП)',
        'ФТГ']
    code_names_columns_ATH = ['Анатомический орган или система', 'Терапевтическая группа', 
       'Фармакологическая группа', 'Химическая группа']
    analysis_part = 'ЛП'
    diff_LP_df = []
    df_a = df_LP[lp_mask_base | lp_mask_techno]
    n_bars_max_on_picture = 20
    # colors=["#9b59b6", "#3498db", "#95a5a6", "#e74c3c", "#34495e", "#2ecc71"]
    cmap = ListedColormap(["#95a5a6", "#2ecc71"])
    y_lim_min = 0
    
    for i_col, col_name in enumerate(columns_to_compare):
        df_p = pd.DataFrame({'count' : df_a.groupby( ['Файл Excel', col_name] ).size()}).reset_index().pivot([col_name], ['Файл Excel'] )\
        .fillna(0)
        # display(df_p.reset_index())
        kind = 'bar' 
        df_pp = simplify_multi_index (df_p, tk_names, model_names)
        # print("df_pp.shape[0]:", df_pp.shape[0])
        kind = 'bar' #'kde' #'area' #'bar'
        title = '\n'.join([tk_code_name, analysis_part])
        y_lim_min = 0
        if df_pp.shape[0] <= n_bars_max_on_picture:
            plt.figure(figsize=(25, 6), tight_layout=True)
            ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)
        else:
            plt.figure(figsize=(25, 10), tight_layout=True)
            for i_max in range(10):
                # df_pp1 = df_pp[(df_pp['База']>=y_lim_min + i_max) | (df_pp['Техно']>=y_lim_min + i_max)]
                df_pp1 = df_pp[(df_pp[model_names[0]]>=y_lim_min + i_max) | (df_pp[model_names[1]]>=y_lim_min + i_max)]
                if df_pp1.shape[0] <= n_bars_max_on_picture:
                    ax1 = df_pp1.plot(kind= kind, x = col_name, rot=45, cmap = cmap) 
                    break

        legend_list = model_names
        ax1.legend(legend_list, loc='best',fontsize=8)
        plt.title(title, fontsize=8)
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)
        plt.xlabel(col_name, fontsize=8)
        plt.ylabel('Количество', fontsize=8)

        # fn_img = f"{analysis_part_code:02d}_{analysis_part}_{i_col:02d}.jpg"
        fn_img = f"{analysis_part_code:02d}_ЛП_01_Состав_{i_col:02d}.jpg" #.replace(' ','_')
        # plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, fn_img), bbox_inches='tight')
        plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, 'img', fn_img), bbox_inches='tight')
        plt.show()
        
        if col_name in columns_to_compare[0:4]:
            diff_LP_df.append(def_differencies(
                df_pp, tk_names, model_names, code_names_columns = code_names_columns_ATH, function_extract_names=extract_name_groups_ATH))
        else: 
            diff_LP_df.append(def_differencies(df_pp, tk_names, model_names, [], function_extract_names=None))
        
        # diff_df_LP.append(def_differencies(
        #                      df_pp, tk_names, model_names,
        #                      code_names_columns = code_names_columns_services, 
        #                      function_extract_names = extract_names_from_code_service))
        display(diff_LP_df[i_col])
    return diff_LP_df

def LP_analysis_02(
    df_LP, tk_names, model_names, tk_code_name,
    path_tk_models_processed,
    analysis_subpart_code, analysis_subpart,
    indicator_col_name = 'Усредненная частота предоставления',
    agg_type = 'Среднее',
    ):
    # print("analysis_part:", analysis_part)
    lp_mask_base = df_LP['Файл Excel'] == tk_names[0]
    lp_mask_techno = df_LP['Файл Excel'] == tk_names[1]
    
    # tk_name, analysis_part, analysis_part_code = tk_code_name, 'ЛП', 2
    columns_to_compare =['Код анатомического органа или системы',
       'Код терапевтической группы', 
       'Код фармакологической группы',
       'Код химической группы', 
       'Код группы ЛП (АТХ)', 
        'Форма выпуска лекарственного препарата (ЛП)',
        'ФТГ']
    code_names_columns_ATH = ['Анатомический орган или система', 'Терапевтическая группа', 
       'Фармакологическая группа', 'Химическая группа']
    # analysis_part = 'ЛП'
    diff_LP_df = []
    df_a = df_LP[lp_mask_base | lp_mask_techno]
    n_bars_max_on_picture = 20
    # colors=["#9b59b6", "#3498db", "#95a5a6", "#e74c3c", "#34495e", "#2ecc71"]
    cmap = ListedColormap(["#95a5a6", "#2ecc71"])
    y_lim_min = 0
    diff_lst = []
    
    for i_col, col_name in enumerate(columns_to_compare):
        diff_lst.append([])
        if agg_type == 'Среднее':
            df_p = df_a.groupby( ['Файл Excel', col_name, ] ).agg({indicator_col_name: ['mean']})\
                        .reset_index().pivot([col_name], ['Файл Excel'] ).fillna(0)
        elif agg_type == 'Сумма':
            df_p = df_a.groupby( ['Файл Excel', col_name, ] ).agg({indicator_col_name: ['sum']})\
                        .reset_index().pivot([col_name], ['Файл Excel'] ).fillna(0)
        # print(df_p.columns)
        # display(df_p.head(2))
        df_pp = simplify_multi_index_02 (df_p, tk_names, model_names)
        # df_pp = simpl_multi_index_02 (df_p, tk_names, model_names)
        # display(df_pp.head(2))
        kind = 'bar' #'kde' #'area' #'bar'
        title = '\n'.join([tk_code_name, 'ЛП', analysis_subpart]) #, indicator_col_name]) #, col_name])
        # print("title:", title)
        y_lim_min = 0
        
        print(f"df_pp.shape[0]:", df_pp.shape[0])
        if df_pp.shape[0] <= n_bars_max_on_picture:
            plt.figure(figsize=(25, 6), tight_layout=True)
            ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)
        else:
            plt.figure(figsize=(25, 10), tight_layout=True)
            try:
                max_v = max(df_pp[model_names[0]].max(), df_pp[model_names[1]].max())
                min_v = min(df_pp[model_names[0]].min(), df_pp[model_names[1]].min())
                delta_v = (max_v - min_v)/10
                print("delta_v:", delta_v)
                if  delta_v != 0:
                    fl_break = False
                    for i_max in range(10):
                        # df_pp1 = df_pp[(df_pp[model_names[0]]>=y_lim_min + i_max*delta_v) | (df_pp[model_names[1]]>=y_lim_min + i_max*delta_v)]
                        df_pp1 = df_pp[(df_pp[model_names[0]]>=min_v + i_max*delta_v) | (df_pp[model_names[1]]>=min_v + i_max*delta_v)]
                        if df_pp1.shape[0] <= n_bars_max_on_picture:
                            print(f"i_max: {i_max}, df_pp1.shape[0]: {df_pp1.shape[0]}")
                            ax1 = df_pp1.plot(kind= kind, x = col_name, rot=45, cmap = cmap) #, y_lim= (y_lim_min + i_max,100)) 
                            fl_break = True
                            break
                    if not fl_break:
                        ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)
                else:
                    ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)
            except Exception as err:
                print(str(err))
                ax1 = df_pp.plot(kind= kind, x = col_name, rot=45, cmap = cmap)

        legend_list = model_names
        ax1.legend(legend_list, loc='best',fontsize=8)
        plt.title(title, fontsize=8)
        plt.xticks(fontsize=8)
        plt.yticks(fontsize=8)
        plt.xlabel(col_name, fontsize=8)
        # plt.ylabel('Количество', fontsize=8)
        plt.ylabel(agg_type, fontsize=8)

        # fn_img = f"{analysis_part_code:02d}_ЛП_{analysis_subpart}_{i_col:02d}.jpg"
        fn_img = f"02_ЛП_{analysis_subpart_code:02d}_{analysis_subpart}_{i_col:02d}.jpg" #.replace(' ','_')
        
        # plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, fn_img), bbox_inches='tight')
        plt.savefig(os.path.join(path_tk_models_processed, tk_code_name, 'img', fn_img), bbox_inches='tight')
        plt.show()
            
        if col_name in columns_to_compare[0:4]:
            try:
                diff_LP_df.append(def_differencies(
                    df_pp, tk_names, model_names, code_names_columns = code_names_columns_ATH, 
                    function_extract_names=extract_name_groups_ATH))
                display(diff_LP_df[i_col])
            except Exception as err:
                diff_LP_df.append(None)
                logger.error(str(err))
                logger.error(f"Данные анализа об отличиях не выводятся из-за некорректных входных данных")
        else: 
            try:
                diff_LP_df.append(def_differencies(df_pp, tk_names, model_names, [], function_extract_names=None))
                display(diff_LP_df[i_col])
            except Exception as err:
                diff_LP_df.append(None)
                logger.error(str(err))
                logger.error(f"Данные анализа об отличиях не выводятся из-за некорректных входных данных")
        
        # diff_df_LP.append(def_differencies(
        #                      df_pp, tk_names, model_names,
        #                      code_names_columns = code_names_columns_services, 
        #                      function_extract_names = extract_names_from_code_service))
        
    return diff_LP_df

def update_excel_by_analysis(
    diff_df_services, diff_LP_df,
    tk_save_dir, tk_code_name, fn_TK_save):
    
    wb = load_workbook(os.path.join(tk_save_dir, tk_code_name, fn_TK_save))
    # tk_name, model, analysis_part, analysis_part_code = 'Нейрохирургия',  'База', 'Услуги', 1
    df_diff = [diff_df_services, diff_LP_df, None]
    cols_width_compare = [[30,70], [30,70], [30,70]]
    cols_width_analysis = [[10,7,7,7,30,30,30], [10,7,7,7,30,30,30,30], None]
    interval_row = 1
    for i_p, analysis_part in enumerate(['Услуги', 'ЛП', 'РМ']):
        sheet_name = analysis_part + '_Сравнение'
        sheet_names = wb.get_sheet_names()
        if sheet_name in sheet_names and cols_width_compare[i_p] is not None:
            ws = wb[sheet_name]
            ws = format_excel_cols_short(ws, cols_width_compare[i_p], auto_filter=True)
    for i_p, analysis_part in enumerate(['Услуги', 'ЛП']): #, 'РМ']):
        # fn_img_lst = glob.glob(os.path.join(tk_save_dir, tk_code_name) + f"{i_p+1:02d}_*.jpg")
        fn_img_lst = glob.glob(os.path.join(
            tk_save_dir, tk_code_name, 'img') + f"/{i_p+1:02d}_{analysis_part}_01_Состав_*.jpg")
            # tk_save_dir, tk_code_name) + f"/{i_p+1:02d}_{analysis_part}_01_Состав_*.jpg")
        print("fn_img_lst:", len(fn_img_lst), fn_img_lst)
        sheet_name = analysis_part + '_Анализ_Состав'
        sheet_names = wb.get_sheet_names()
        if sheet_name in sheet_names:
            # wb.remove_sheet(sheet_name)
            wb.remove(wb[sheet_name])
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]
        if cols_width_analysis[i_p] is not None:
            ws = format_excel_cols_short(ws, cols_width_analysis[i_p], auto_filter=False)
        # cell = ws['A1']
        # font_size = cell.font.sz
        cell_height = 20 # опытным путем
        cell_height = 17 # опытным путем

        images_total_height = 0
        images_total_rows = 0
        explain_rows = 0
        interval_rows = 0

        for i_f, fn_img in enumerate(fn_img_lst):
            img = drawing.image.Image(fn_img)
            anchor = f"A{images_total_rows + explain_rows+1}" 
            ws.add_image(img, anchor)
            # img_rows = int(img.height//cell_height   + 1) # + interval_row
            img_rows = img.height//cell_height   + 1 + 2*interval_row
            images_total_rows += img_rows
            for _ in range(img_rows): ws.append([None])

            if df_diff[i_p] is not None:
                # cell = ws[anchor]
                # print(f"i_p: {i_p}, len(df_diff[i_p]):", len(df_diff[i_p]))
                ws.append(list(df_diff[i_p][i_f].columns))
                for i_row, row in df_diff[i_p][i_f].iterrows():
                    ws.append(list(row.values))
                explain_rows += df_diff[i_p][i_f].shape[0] + 1 + 2*interval_row

            # print(img.height, img_rows, images_total_rows, explain_rows)

    wb.save(os.path.join(tk_save_dir, tk_code_name, fn_TK_save))   
    logger.info(f"Файл '{fn_TK_save}' дополнен данными анализа и сохранен в '{os.path.join(tk_save_dir, tk_code_name, fn_TK_save)}'")



def update_excel_by_analysis_02(
    diff_df_services_02, diff_LP_df_02,
    path_tk_models_processed, tk_code_name, fn_TK_save):
    
    wb = load_workbook(os.path.join(path_tk_models_processed, tk_code_name, fn_TK_save))
    # tk_name, model, analysis_part, analysis_part_code = 'Нейрохирургия',  'База', 'Услуги', 1
    df_diff = [diff_df_services_02, diff_LP_df_02, None]
    cols_width_analysis = [[10,7,7,7,30,30,30], [10,7,7,7,30,30,30,30], None]
    interval_row = 1
    analysis_subpart_lst = [ [(2, 'Частота'), (3, 'Кратность'), (4, 'УЕТ 1'), (5, 'УЕТ 2')],
                             [(2, 'Частота'), (3, 'Кратность'), (4, 'Количество')]
    ]
    for i_p, analysis_part in enumerate(['Услуги', 'ЛП']): #, 'РМ']):
        for i_sp, (analysis_subpart_code, analysis_subpart) in enumerate(analysis_subpart_lst[i_p]):
            fn_img_lst = glob.glob(os.path.join(
                path_tk_models_processed, tk_code_name, 'img') + f"/{i_p+1:02d}_{analysis_part}_{analysis_subpart_code:02d}_{analysis_subpart}*.jpg")
                # path_tk_models_processed, tk_code_name) + f"/{i_p+1:02d}_{analysis_part}_{analysis_subpart_code:02d}_{analysis_subpart}*.jpg")
                # path_tk_models_processed, tk_code_name) + f"/{i_p+1:02d}_{analysis_part}_{i_sp:02d}_{analysis_subpart.replace(' ', '_')}*.jpg")
            print("fn_img_lst:", len(fn_img_lst), fn_img_lst)
            sheet_name = analysis_part + '_Анализ_' + analysis_subpart #.replace(' ', '_')
            sheet_names = wb.get_sheet_names()
            if sheet_name in sheet_names:
                # wb.remove_sheet(sheet_name)
                wb.remove(wb[sheet_name])
            wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            if cols_width_analysis[i_p] is not None:
                ws = format_excel_cols_short(ws, cols_width_analysis[i_p], auto_filter=False)
            # cell = ws['A1']
            # font_size = cell.font.sz
            cell_height = 20 # опытным путем
            cell_height = 17 # опытным путем

            images_total_height = 0
            images_total_rows = 0
            explain_rows = 0
            interval_rows = 0

            for i_f, fn_img in enumerate(fn_img_lst):
                img = drawing.image.Image(fn_img)
                anchor = f"A{images_total_rows + explain_rows+1}" 
                ws.add_image(img, anchor)
                # img_rows = int(img.height//cell_height   + 1) # + interval_row
                img_rows = img.height//cell_height   + 1 + 2*interval_row
                images_total_rows += img_rows
                for _ in range(img_rows): ws.append([None])

                if df_diff[i_p] is not None:
                    # cell = ws[anchor]
                    # print(f"i_p: {i_p}, len(df_diff[i_p]):", len(df_diff[i_p]))
                    if df_diff[i_p][i_sp] is not None:
                    # if df_diff[i_p][i_f] is not None:
                        try:
                            # ws.append(list(df_diff[i_p][i_f].columns))
                            # for i_row, row in df_diff[i_p][i_f].iterrows():
                            ws.append(list(df_diff[i_p][i_sp][i_f].columns))
                            for i_row, row in df_diff[i_p][i_sp][i_f].iterrows():
                                ws.append(list(row.values))
                            explain_rows += df_diff[i_p][i_sp][i_f].shape[0] + 1 + 2*interval_row
                        except Exception as err:
                            print(err)
                            # print(type(df_diff[i_p][i_f]), df_diff[i_p][i_f])
                    else:
                        for i_row, row in range(2*interval_row):
                            ws.append([None])
                        explain_rows += 2*interval_row

            # print(img.height, img_rows, images_total_rows, explain_rows)

    wb.save(os.path.join(path_tk_models_processed, tk_code_name, fn_TK_save))   
    logger.info(f"Файл '{fn_TK_save}' дополнен данными анализа и сохранен в '{os.path.join(path_tk_models_processed, tk_code_name)}'")


def data_comparsion (
    df_services, df_LP, df_RM,
    tk_code_name, tk_names, model_names,
    path_tk_models_processed
):
    df_services_compare = services_comparison(
        df_services, tk_names, model_names,
        col_to_compare = 'Наименование услуги по Номенклатуре медицинских услуг (Приказ МЗ №804н)')    
    df_LP_compare = LP_comparison(
        df_LP, tk_names, model_names,
        col_to_compare = 'Наименование лекарственного препарата (ЛП) (МНН)')    
    df_RM_compare = RM_comparison(
        df_RM, tk_names, model_names,
        col_to_compare = 'Изделия медицинского назначения и расходные материалы, обязательно используемые при оказании медицинской услуги')

    tk_save_dir = os.path.join(path_tk_models_processed, tk_code_name)
    if not os.path.exists(tk_save_dir): os.mkdir(tk_save_dir)
    fn_TK_save = save_df_lst_to_excel([df_services_compare, df_LP_compare, df_RM_compare], ['Услуги_Сравнение', 'ЛП_Сравнение', 'РМ_Сравнение'], tk_save_dir, tk_code_name + '.xlsx')
    logger.info(f"Файл '{fn_TK_save}' сохранен в директорию '{tk_save_dir}'")
    
    return fn_TK_save

def data_analysis_composition(
    df_services, df_LP,
    tk_names, model_names, tk_code_name,
    path_tk_models_processed, fn_TK_save
):
    diff_df_services = services_analysis(
        df_services, 
        tk_names, model_names, tk_code_name,
        path_tk_models_processed
        )
    diff_LP_df = LP_analysis(
        df_LP, 
        tk_names, model_names, tk_code_name,
        path_tk_models_processed
    )   
    update_excel_by_analysis(diff_df_services, diff_LP_df, path_tk_models_processed, tk_code_name, fn_TK_save )

def data_analysis_02(
    df_services, df_LP,
    tk_names, model_names, tk_code_name,
    path_tk_models_processed, fn_TK_save
):
    service_indicators_lst = [(2, 'Частота', 'Усредненная частота предоставления', 'Среднее'),
                           (3, 'Кратность', 'Усредненная кратность применения', 'Среднее'),
                           (4, 'УЕТ 1', 'УЕТ 1', 'Сумма'), 
                           (5, 'УЕТ 2', 'УЕТ 2', 'Сумма')]
    LP_indicators_lst = [(2, 'Частота', 'Усредненная частота предоставления', 'Среднее'),
                           (3, 'Кратность', 'Усредненная кратность применения', 'Среднее'),
                           (4, 'Количество', 'Кол-во', 'Сумма'), 
                           ]
    diff_df_services_02 = []
    diff_LP_df_02 = []
    for analysis_subpart_code, analysis_subpart, indicator_col_name, agg_type in service_indicators_lst:
        diff_df_services_02.append(services_analysis_02(
                                        df_services, tk_names, model_names, tk_code_name,
                                        path_tk_models_processed,
                                        analysis_subpart_code, analysis_subpart,
                                        indicator_col_name = indicator_col_name,
                                        agg_type = agg_type,
                                                        )
                                  )
    print("len(diff_df_services_02):",len(diff_df_services_02))
    # print(diff_df_services_02)
    for analysis_subpart_code, analysis_subpart, indicator_col_name, agg_type in LP_indicators_lst:
        diff_LP_df_02.append(LP_analysis_02(
                                        df_LP, tk_names, model_names, tk_code_name,
                                        path_tk_models_processed,
                                        analysis_subpart_code, analysis_subpart,
                                        indicator_col_name = indicator_col_name,
                                        agg_type = agg_type,
                                        )     
                            )
        # update_excel_by_analysis(diff_df_services, diff_LP_df, path_tk_models_processed, tk_code_name, fn_TK_save )
    update_excel_by_analysis_02(diff_df_services_02, diff_LP_df_02, path_tk_models_processed, tk_code_name, fn_TK_save )
    
    return diff_df_services_02, diff_LP_df_02




def total_comparsion_analysis(
    path_tk_models_source, fn_tk_description,
    df_services, df_LP, df_RM,
    path_tk_models_processed,
    # first_model = 'База'
):
    tk_models = read_description(path_tk_models_source, fn_tk_description)
    
    for tk_full_name, tk_dict in tk_models.items():
        tk_name_short = tk_dict.get('Наименование ТК (короткое)')
        tk_code = tk_dict.get('Код ТК')
        tk_code_name = str(tk_code) + '_' + tk_name_short
        if tk_dict.get('Модели') is not None and len(tk_dict.get('Модели')) > 1:
            models_dict_lst = tk_dict.get('Модели')
            models_dict_lst.sort(key = lambda model_dict: model_dict['Модель пациента'])
            tk_names = [models_dict_lst[0]['Файл Excel'], models_dict_lst[1]['Файл Excel'] ]
            model_names = [models_dict_lst[0]['Модель пациента'], models_dict_lst[1]['Модель пациента'] ]
            print(tk_names)
            print(model_names)
            fn_TK_save = data_comparsion (
                    df_services, df_LP, df_RM,
                    tk_code_name, tk_names, model_names,
                    path_tk_models_processed
            )
            data_analysis_composition(
                df_services, df_LP,
                tk_names, model_names, tk_code_name,
                path_tk_models_processed, fn_TK_save    
            )
            
            data_analysis_02(
                df_services, df_LP,
                tk_names, model_names, tk_code_name,
                path_tk_models_processed, fn_TK_save    
            )

def compare_tk( data_source_dir, data_processed_dir, supp_dict_dir, fn_tk_bd, fn_tk_description):

    if fn_tk_bd is None or  fn_tk_description is None:
        logger.error(f"Выберите названия файлов: сводного и описания моделей - в параметрах запуска программы")
        sys.exit(2)
    df_services, df_LP, df_RM = preprocess_tkbd(data_source_dir, fn_tk_bd, data_processed_dir, supp_dict_dir)
    display(df_services.head(2))
    display(df_LP.head(2))
    display(df_RM.head(2))
    # path_tk_models_source = data_source_dir
    # print(f"path_tk_models_source: '{path_tk_models_source}'")
    total_comparsion_analysis(data_source_dir, fn_tk_description, df_services, df_LP, df_RM, data_processed_dir)
